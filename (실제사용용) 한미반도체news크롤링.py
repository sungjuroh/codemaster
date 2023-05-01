from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time


# News 크롤링 할 Excel 불러오기
import openpyxl
import datetime

d_today = datetime.date.today()  # 오늘 날짜 추출위한
fpath = r'C:\Users\sgn31\OneDrive\바탕 화면\No\Programming\selenium\한미반도체_News.xlsx'   # 파일 위치 경로는 따로 변수에 저장

wb = openpyxl.load_workbook(fpath)
ws = wb.create_sheet(str(d_today))   # 여기까지 하면 Excel 생성 및 오늘 날짜 Sheet 생성까지 완료


# URL 크롤링 하기

def crawling_img(name):
    driver = webdriver.Chrome()
    driver.get("https://news.naver.com")
    elem = driver.find_element(By.CSS_SELECTOR,".Nicon_search")
    elem.click()
    elem = driver.find_element(By.NAME,"query")
    elem.send_keys(name)
    elem.send_keys(Keys.RETURN)
    driver.switch_to.window(driver.window_handles[-1])

    #
    SCROLL_PAUSE_TIME = 1
    # Get scroll height
    last_height = driver.execute_script("return document.body.scrollHeight")  # 브라우저의 높이를 자바스크립트로 찾음
    while True:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")  # 브라우저 끝까지 스크롤을 내림
        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)
        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
      

    urls = driver.find_elements(By.CSS_SELECTOR, '.news_tit')   # 화면 上 제목들의 URL...

    count = 1
    for url in urls:
        try:
            # time.sleep(5)            
            ws[f'B{2*count}'] = str(count)+'.'                         # 활성 Sheet의 B 짝수열에는 "번호." 형식
            ws[f'C{2*count}'] = url.get_attribute("href")              # 활성 Sheet의 C 짝수열에는 "URL" 형식
            count = count + 1
            if count >= 10:
                break
        except:
            pass



    wb.save(r'C:\Users\sgn31\OneDrive\바탕 화면\No\Programming\selenium\한미반도체_News.xlsx')    # 저장


    driver.close()
idols = ["한미반도체"]

for idol in idols:
    crawling_img(idol)
