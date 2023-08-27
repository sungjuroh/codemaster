from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import datetime
import pyperclip
import pyautogui
import undetected_chromedriver as uc
import subprocess
import os
import openpyxl
from openpyxl import load_workbook
import warnings
import pandas as pd
import sqlite3
from sqlite3 import Error
import logging
import win32com.client
import requests

logger = logging.getLogger("my_log")
logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()  # 콘솔 표시
ch.setLevel(logging.DEBUG)
fh = logging.FileHandler("my_log.log")  # 로그 파일에 저장
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s, %(name)s, line : %(lineno)d, %(levelname)s : %(message)s',
                              datefmt='%Y-%m-%d %H:%M:%S %p')
ch.setFormatter(formatter)
logger.addHandler(ch)
fh.setFormatter(formatter)
logger.addHandler(fh)

# logger.debug('실행완료')  
# logger.info('진행중...')  
# logger.warning('문제가 있는 것 같습니다..')
# logger.error('에러가 발생했습니다.')
# logger.critical('종료합니다.')


warnings.simplefilter("ignore")

subprocess.Popen(r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\Users\sgn31\OneDrive\바탕 화면\No\Programming\selenium\디버그 모드"')
url = "https://kbland.kr/"
option = Options()
option.add_argument("--headless")                                  # chrome 창 안띄우고 진행되도록 옵션
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
# release = "https://chromedriver.storage.googleapis.com/LATEST_RELEASE"
# version = requests.get(release).text
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=option)
driver.maximize_window()
driver.get(url)

# 로그인 방식 클릭
time.sleep(2)
aaa = driver.find_element(By.XPATH,'//*[@id="app"]/div/div[7]/div/div[1]/button')
aaa.click()
time.sleep(2)
aaa = driver.find_element(By.XPATH,'//*[@id="app"]/div/div[9]/div[2]/div/div/div[1]/div/section/div/div/div[3]/button')
aaa.click()
time.sleep(2)
logger.info('ing...')  

# 구글로그인 클릭
aaa = driver.execute_script('document.querySelector("#app > div > div.memberArea.active > div.memberContent > div > div > div.scrollbar-inner.scroll-content > section > div.loginCont > div.btns > button.btn.btn-login.google > span").click();',aaa)
time.sleep(5)
logger.debug('Google Log-in complete')

# 지도 검색창으로 이동
aaa = driver.find_element(By.XPATH,'//*[@id="app"]/div/nav/div/button[2]')
aaa.click()
time.sleep(3)

# 검색창 누르기
aaa = driver.find_element(By.XPATH,'//*[@id="searchArea"]/div/button[1]')
aaa.click()
time.sleep(3)
logger.info('map searching...') 

# 대상 부동산 입력 및 검색
pyautogui.press('Tab',9)
aaa.send_keys('경기도 화성시 동탄순환대로 881')
pyautogui.press('enter')
logger.info('apart searching...') 
time.sleep(3)
logger.debug('searching complete')

# 시세 Excel 다운로드
aaa = driver.find_element(By.XPATH,'//*[@id="app"]/div/div[8]/div/div/span')
aaa.click()
time.sleep(3)
pyperclip.copy('KB시세 다운로드') # 클립보드에 텍스트를 복사합니다. 
pyautogui.hotkey('ctrl', 'f')
pyautogui.hotkey('ctrl', 'v')
pyautogui.press('enter')
time.sleep(3)
aaa = driver.find_element(By.XPATH,'//*[@id="시세"]/div[2]/div/div[3]/button[2]')  # KB시세 다운로드 클릭
aaa.click()
time.sleep(2)
aaa = driver.find_element(By.XPATH,'//*[@id="시세"]/div[2]/div/div[4]/div[2]/div/div/div[2]/button[1]/strong/em')  # 과거 시세 다운로드 클릭
aaa.click()
logger.info('apart price Excel Download...') 
time.sleep(3)
logger.debug('Download complete')

# 다운로드 파일 이름 변경 및 관리 폴더 쪽으로 위치 이동
d_today = datetime.date.today()

download_path = r"C:\Users\Public"
Transfer_path = r"C:\Users\Public\부동산시세\동탄역푸르지오"

filename = "동탄역푸르지오 과거시세.xlsx"
new_filename = "(" + str(d_today) + ")" + filename

old_path = os.path.join(download_path, filename)
new_path = os.path.join(Transfer_path, new_filename)

os.rename(old_path, new_path)

time.sleep(1)
logger.debug('apart_file chdir complete')

driver.close()
logger.debug('Chrome close')

wb = openpyxl.load_workbook(new_path, data_only=True)
ws = wb.active                   # (날짜)동탄역푸르지오 엑셀 파일 실행
logger.debug('apart excel file execute')

data_dict_list = []

def extract() :
    # global data #https://wikidocs.net/24#2-global
    # global data_dict
    for j in range(16,89):
        date = ws['A'+str(j)].value
        lower_value = ws['B'+str(j)].value
        mean_value = ws['C'+str(j)].value
        upper_value = ws['D'+str(j)].value
        # if문 사용해서 cell 내 데이터 없으면 넘어가는 코드 생성 필요
        # print([Upper_value,Mean_value,Lower_value])
        data = {'date' : date, 'lower' : lower_value, 'mean' : mean_value, 'upper' : upper_value} # https://wikidocs.net/16#_2
        data_dict_list.append(data)

    return data_dict_list

data_dict_list = extract() #함수의 선언과 호출 (점프 투 파이썬에는 설명이 빈약함. 추가 이미지 첨부) https://wikidocs.net/24#1-return

logger.debug('Data from excel extract completed')

os.chdir('C:/Users/Public/부동산시세')

def connection() :
    try:
        con = sqlite3.connect('data_apart_동탄역푸르지오.db')
        return con
    except Error:
        print(Error)


def create_table(con):
    cursor_db = con.cursor()
    cursor_db.execute('CREATE TABLE IF NOT EXISTS roh(id integer PRIMARY KEY, date date, lower real, mean real, upper real)')
    con.commit()

def insert_one(con, one_data):
    cursor_db = con.cursor()
    cursor_db.execute('INSERT OR REPLACE INTO roh(id, date, lower, mean, upper) VALUES(?, ?, ?, ?, ?)', one_data)
    con.commit()

con = connection()

create_table(con)
for i in range(len(data_dict_list)): #점프투 파이썬 02-3 리스트 자료형 3.3.리스트 자료형 구하기 https://wikidocs.net/14#_8
    one_data = (i, data_dict_list[i]['date'], data_dict_list[i]['lower'], data_dict_list[i]['mean'], data_dict_list[i]['upper']) #딕셔너리 리스트 호출 책에 내용이 없어서 자료 첨부  
    insert_one(con, one_data)

logger.debug('Data insert to DB complete')

# 요 부분은 확인 필요..DB 자료 변동 안됨
cursor_db = con.cursor()
cursor_db.execute('select * from roh order by date asc')
con.commit()


wb.close

logger.debug('DB saved, excel close')                                  


# ws.Range("A1:D100").Copy() # DB로 해당 셀 내용들 보내기

# wd.Range("A1").Select()     
# wd.Paste()                  # 부동산크롤링 엑셀파일 A1 선택 & 붙여넣기

# for i in range(1,16):
#     wd.Rows(1).EntireRow.Delete()

# wd.Columns("A:D").Sort(Key1=wd.Range("A1"),Order1=1, Orientation=1)
# wd.Rows(1).EntireRow.Insert()

# # time.sleep(2)

# wd.Range("A1").value = "동탄역푸르지오"
# wd.Range("B1").value = "하위 평균가"
# wd.Range("C1").value = "일반 평균가"
# wd.Range("D1").value = "상위 평균가"

# wb.save
# wc.save

# wb.close
# wc.close

# driver.close()

# wf = openpyxl.load_workbook(apart_path)
# wg = wf['Sheet1']

# chart = openpyxl.chart.LineChart()
# chart.title = "동탄역푸르지오 시세 추이"
# chart.x_axis.title = "시세기준월"
# chart.y_axis.title = "시세(만원)"
# datas = openpyxl.chart.Reference(wg, min_col=2, min_row=1, max_col=4, max_row = 85)
# chart.add_data(datas, from_rows=False, titles_from_data=True)
# cats = openpyxl.chart.Reference(wg, min_col=1, min_row=2, max_col=1, max_row = 85)
# chart.set_categories(cats)
# chart.height = 15
# chart.width = 30
# wg.add_chart(chart, "F2")
# wf.save(apart_path)
# wf.close


